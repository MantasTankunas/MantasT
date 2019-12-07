import openpyxl as xl

class Gaminys:

    def __init__(self,Prefix, Name, Volume, Total_Volume, Width, Height, Lenght, Reinforcement, Floor):
        self.Prefix = Prefix
        self.Name = Name
        self.Volume = Volume
        self.Total_Volume = Total_Volume
        self.Width = Width
        self.Height = Height
        self.Lenght = Lenght
        self.Reinforcement = Reinforcement
        self.Floor = Floor

def istraukti_duomenis(filename=r'C:\Users\MantasT\Desktop\uzduotis_e\Isranka_galutine.xlsx'):
    gaminiai = []
    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]
    mr = ws1.max_row  # numeris paskutines eilutes
    mc = ws1.max_column  # numeris paskutinio stulpelio
    for i in range(1, mr + 1):  # suka nuo 1 iki paskutines eilutes +1
        for j in range(1, mc + 1):  # suka nuo 1 iki paskutinio stulpelio +1
            # reading cell value from source excel file
            c = ws1.cell(row=i, column=j).value
            print(c)
        # writing the read value to destination excel file
#            ws2.cell(row=i, column=j).value = c.value

    g = Gaminys('VS-151','Name', 'Volume', 'Total_Volume', 'Width', 'Height', 'Lenght', 'Reinforcement', 'Floor')
    gaminiai.append(g)

    # saving the destination excel file
    #        wb2.save(str(filename))
    return gaminiai


print(istraukti_duomenis(filename=r'C:\Users\MantasT\Desktop\uzduotis_e\Isranka.xlsx'))

# gaminys = gaminiai(VS-151Internal wall		0,56
# )


